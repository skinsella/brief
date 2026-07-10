"""
Cross-validate our LLM AI-exposure scores against the academic Felten,
Raj & Seamans (2021) AIOE benchmark.

Felten/Raj/Seamans construct an AI Occupational Exposure (AIOE) score
for ~774 US SOC 2018 6-digit occupations using O*NET ability ratings
weighted by AI application–ability matches. Values are z-scored
(roughly −3 to +3); higher = more exposed. The dataset is in
raw/aioe_appendix.xlsx (downloaded from github.com/AIOE-Data/AIOE).

This script:
  1. Loads AIOE Appendix A (SOC code, title, score).
  2. For each NSB occupational group, looks at its list of UK SOC2010
     description strings (already in occupations.json), token-matches
     each to the closest US SOC 2018 title in AIOE, and averages the
     AIOE scores.
  3. Normalises AIOE to a 0–10 scale comparable with our LLM scores.
  4. Writes raw/_aioe_benchmark.json keyed by occupation slug, with
     fields {aioe_raw, aioe_0_10, n_matched, matched_socs}.

Run: uv run python scripts/07_aioe_benchmark.py
"""

from __future__ import annotations

import json
import re
import statistics
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
INDEX_IN = REPO / "occupations.json"
AIOE_XLSX = REPO / "raw" / "aioe_appendix.xlsx"
OUT = REPO / "raw" / "_aioe_benchmark.json"

STOPWORDS = {
    "and", "the", "of", "in", "for", "etc", "n", "e", "c", "&",
    "a", "an", "or", "with", "to", "all", "other", "n.e.c.",
}


def toks(s: str) -> set[str]:
    return {t for t in re.findall(r"\w+", s.lower()) if t not in STOPWORDS}


def load_aioe_rows() -> list[dict]:
    """Return [{soc, title, aioe, tokens, is_catchall}] from Appendix A,
    with an IDF score for each token across the whole AIOE corpus.
    """
    wb = load_workbook(AIOE_XLSX, data_only=True)
    ws = wb["Appendix A"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        soc, title, aioe = r[0], r[1], r[2]
        if soc and title and aioe is not None:
            t = str(title).strip()
            rows.append({
                "soc": str(soc).strip(),
                "title": t,
                "aioe": float(aioe),
                "tokens": toks(t),
                # Penalise SOC "catch-all" categories like "Engineers, All Other"
                # which match almost anything with a shared common noun.
                "is_catchall": "all other" in t.lower(),
            })
    # Compute IDF for every token: log(N / df_t).
    import math
    from collections import Counter
    df = Counter()
    for row in rows:
        for t in row["tokens"]:
            df[t] += 1
    N = len(rows)
    idf = {t: math.log((N + 1) / (df_t + 1)) + 1 for t, df_t in df.items()}
    # Stash IDF on each row's tokens.
    for row in rows:
        row["idf"] = {t: idf.get(t, 1.0) for t in row["tokens"]}
    # Expose IDF dict at module scope for the matcher.
    global TOKEN_IDF
    TOKEN_IDF = idf
    return rows


TOKEN_IDF: dict[str, float] = {}


def best_aioe_match(soc_desc: str, aioe_rows: list[dict]) -> dict | None:
    """Find the AIOE row whose title best matches a UK SOC2010 description
    string (token overlap + stem-aware scoring). Returns the row dict or
    None if no plausible match found.
    """
    desc_toks = toks(soc_desc)
    if not desc_toks:
        return None
    best = None
    best_score = 0.0
    for row in aioe_rows:
        rt = row["tokens"]
        if not rt:
            continue
        common = desc_toks & rt
        # IDF-weighted overlap: rare tokens (e.g. "plumbers", "actuaries")
        # dominate; common tokens (e.g. "engineers", "managers") barely
        # contribute. This handles the "All Other" catch-all problem
        # without explicit exclusion.
        match_w = sum(TOKEN_IDF.get(t, 1.0) for t in common)
        total_w = sum(TOKEN_IDF.get(t, 1.0) for t in desc_toks | rt) or 1.0
        score = match_w / total_w
        # Stem-prefix bonus for distinct 5-char prefixes.
        for a in desc_toks:
            for b in rt:
                if a == b:
                    continue
                if len(a) >= 6 and len(b) >= 6 and a[:5] == b[:5]:
                    score += 0.05 * TOKEN_IDF.get(b, 1.0) / 4
        if row.get("is_catchall"):
            score *= 0.5  # halve catch-all "All Other" entries
        if score > best_score:
            best_score = score
            best = row

    if best is None:
        return None
    common = desc_toks & best["tokens"]
    # Threshold ~0.18 in normalised IDF-weighted units. Tuned to allow
    # legitimate 1-token matches with a rare token (e.g. "plumbers") while
    # rejecting 1-token matches on common tokens (e.g. "managers").
    if best_score >= 0.18 or len(common) >= 2:
        return best
    return None


def main() -> None:
    entries = json.loads(INDEX_IN.read_text())
    aioe_rows = load_aioe_rows()
    print(f"AIOE Appendix A: {len(aioe_rows)} US SOC 6-digit occupations")
    raw_scores = [r["aioe"] for r in aioe_rows]
    print(f"  AIOE range: {min(raw_scores):.2f} … {max(raw_scores):.2f}")
    print(f"  AIOE mean: {statistics.mean(raw_scores):.3f}, "
          f"stdev: {statistics.stdev(raw_scores):.3f}")

    # Pick scaling so that AIOE ≈ [-3, +3] maps to [0, 10] linearly.
    AIOE_MIN, AIOE_MAX = -3.0, 3.0

    def aioe_to_0_10(v: float) -> float:
        v = max(AIOE_MIN, min(AIOE_MAX, v))
        return round((v - AIOE_MIN) / (AIOE_MAX - AIOE_MIN) * 10, 2)

    out: dict[str, dict] = {}
    matched_n = 0
    for entry in entries:
        slug = entry["slug"]
        socs = entry.get("soc_descriptions", [])
        per_soc: list[dict] = []
        for soc_desc in socs:
            best = best_aioe_match(soc_desc, aioe_rows)
            if best:
                per_soc.append({
                    "us_soc": best["soc"],
                    "us_title": best["title"],
                    "uk_desc": soc_desc,
                    "aioe": best["aioe"],
                })
        if per_soc:
            mean_aioe = statistics.mean(r["aioe"] for r in per_soc)
            out[slug] = {
                "aioe_raw": round(mean_aioe, 4),
                "aioe_0_10": aioe_to_0_10(mean_aioe),
                "n_matched": len(per_soc),
                "n_uk_socs": len(socs),
                "matched": per_soc,
            }
            matched_n += 1

    OUT.write_text(json.dumps(out, indent=2))
    print(f"\nMatched {matched_n}/{len(entries)} NSB groups to AIOE benchmark")
    print(f"Wrote {OUT.relative_to(REPO)}")

    # Distribution check
    vals = [v["aioe_0_10"] for v in out.values()]
    if vals:
        print(f"\nAIOE (0–10 scale) for matched groups:")
        print(f"  mean: {statistics.mean(vals):.2f}")
        print(f"  range: {min(vals):.1f} … {max(vals):.1f}")

    # Sample matches for sanity
    print("\nSample matches:")
    for slug in ["programmers-software-developers", "plumbers",
                 "customer-service-occupations", "nurses-midwives",
                 "accountants-tax-experts"]:
        info = out.get(slug)
        if info:
            print(f"  {slug}: AIOE {info['aioe_raw']:+.2f} → {info['aioe_0_10']}/10 "
                  f"({info['n_matched']}/{info['n_uk_socs']} SOCs matched)")
            for m in info["matched"][:2]:
                print(f"    {m['uk_desc']!r:50s} → {m['us_title']!r} "
                      f"({m['us_soc']}, AIOE {m['aioe']:+.2f})")


if __name__ == "__main__":
    main()

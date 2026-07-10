#!/usr/bin/env python3
"""Fetch the latest Analytical Exchequer Statement from gov.ie.

The Department of Finance publishes the AES monthly (~2nd working day) at a
predictable URL: /en/department-of-finance/publications/
analytical-exchequer-statement-<month>-<year>/, with a single .xlsx asset.
The Databank open-data CSV rotted in 2022, so this parses the Excel directly.

Walks back from the current month until a publication resolves, parses the
cumulative outturn/prior-year/y-o-y columns, and merges the month into
docs/economy/data/exchequer.json (history accumulates across runs, so the
year-to-date series builds month by month).
"""

import json
import ssl
import sys
import urllib.request
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent.parent
OUT = REPO / "docs" / "economy" / "data" / "exchequer.json"

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")
PAGE_TMPL = ("https://www.gov.ie/en/department-of-finance/publications/"
             "analytical-exchequer-statement-{month}-{year}/")

MONTHS = ["january", "february", "march", "april", "may", "june",
          "july", "august", "september", "october", "november", "december"]

# Aggregates match the whole label exactly — prefix matching would catch
# sub-lines like "Expenditure of SIF" before the real (B) Expenditure row.
ROW_KEYS_EXACT = {
    "revenue": "total_revenue",          # (A)
    "expenditure": "total_expenditure",  # (B)
    "tax revenue": "tax_revenue",
    "exchequer balance": "exchequer_balance",
}
# Tax heads are safe to match by prefix (labels carry footnote markers).
ROW_KEYS_PREFIX = {
    "income tax": "income_tax",
    "vat": "vat",
    "excise duties": "excise",
    "corporation tax": "corporation_tax",
    "stamp duties": "stamp_duties",
    "capital gains tax": "cgt",
    "capital acquisitions tax": "cat",
    "customs": "customs",
    "motor tax": "motor_tax",
    "interest on national debt": "debt_interest",
}


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": UA,
                                               "Referer": "https://www.gov.ie/"})
    ctx = ssl.create_default_context()
    with urllib.request.urlopen(req, timeout=60, context=ctx) as r:
        return r.read()


def find_xlsx_url(page_html):
    # first assets.gov.ie …xlsx link on the publication page
    marker = "assets.gov.ie"
    for chunk in page_html.split('href="')[1:]:
        url = chunk.split('"', 1)[0]
        if marker in url and url.lower().endswith(".xlsx"):
            return url
    return None


def discover_latest():
    """Try current month backwards; the AES for month M appears early in M+1."""
    now = datetime.now(timezone.utc)
    year, month = now.year, now.month
    for _ in range(8):
        url = PAGE_TMPL.format(month=MONTHS[month - 1], year=year)
        try:
            html = fetch(url).decode("utf-8", "replace")
            xlsx = find_xlsx_url(html)
            if xlsx:
                return MONTHS[month - 1], year, xlsx
        except Exception:
            pass
        month -= 1
        if month == 0:
            month, year = 12, year - 1
    return None, None, None


def clean_label(v):
    return str(v or "").strip().strip(":-–* ").strip().lower()


def parse_workbook(data):
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for row in ws.iter_rows(max_col=6, values_only=True):
        label = clean_label(row[1])
        key = ROW_KEYS_EXACT.get(label)
        if key is None:
            for prefix, k in ROW_KEYS_PREFIX.items():
                if label.startswith(prefix):
                    key = k
                    break
        if key is None or key in out:
            continue
        try:
            outturn = float(row[2])
            prior = float(row[3])
        except (TypeError, ValueError):
            continue
        yoy_pct = None
        try:
            yoy_pct = round(float(row[5]) * 100, 2)
        except (TypeError, ValueError):
            pass
        out[key] = {
            "outturn_m": round(outturn, 1),
            "prior_year_m": round(prior, 1),
            "yoy_m": round(outturn - prior, 1),
            "yoy_pct": yoy_pct,
        }
    return out


def main():
    month_name, year, xlsx_url = discover_latest()
    if not xlsx_url:
        print("No Analytical Exchequer Statement found in the last 8 months.",
              file=sys.stderr)
        # leave the existing JSON untouched; this is a soft failure
        sys.exit(0 if OUT.exists() else 1)

    print(f"Found: {month_name} {year} → {xlsx_url}")
    figures = parse_workbook(fetch(xlsx_url))
    if "tax_revenue" not in figures:
        print("Parse failed: no tax revenue row found.", file=sys.stderr)
        sys.exit(1)

    month_iso = f"{year}-{MONTHS.index(month_name) + 1:02d}"
    existing = {"history": []}
    if OUT.exists():
        try:
            existing = json.load(open(OUT))
        except (json.JSONDecodeError, OSError):
            pass

    history = [h for h in existing.get("history", []) if h.get("month") != month_iso]
    history.append({"month": month_iso, "source": xlsx_url, **figures})
    history.sort(key=lambda h: h["month"])

    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "latest_month": month_iso,
        "latest": figures,
        "history": history,
        "note": "Cumulative year-to-date figures in €m from the Analytical "
                "Exchequer Statement, Department of Finance.",
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    json.dump(payload, open(OUT, "w"), indent=1)
    tax = figures["tax_revenue"]
    print(f"Wrote {OUT}: to end-{month_name.title()} tax €{tax['outturn_m']/1000:.1f}bn "
          f"({tax['yoy_pct']:+.1f}% y/y), {len(history)} month(s) of history")


if __name__ == "__main__":
    main()
